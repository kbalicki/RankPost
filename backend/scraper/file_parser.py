import io


def parse_txt(content: bytes) -> str:
    for encoding in ("utf-8", "cp1250", "latin-1"):
        try:
            return content.decode(encoding)
        except UnicodeDecodeError:
            continue
    return content.decode("utf-8", errors="replace")


def parse_docx(content: bytes) -> str:
    from docx import Document
    doc = Document(io.BytesIO(content))
    return "\n".join(p.text for p in doc.paragraphs if p.text.strip())


def parse_pdf(content: bytes) -> str:
    from PyPDF2 import PdfReader
    reader = PdfReader(io.BytesIO(content))
    texts = []
    for page in reader.pages:
        text = page.extract_text()
        if text:
            texts.append(text)
    return "\n".join(texts)


def parse_xlsx(content: bytes) -> str:
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    lines = []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) if c is not None else "" for c in row]
            line = " | ".join(cells).strip()
            if line and line != "| " * len(cells):
                lines.append(line)
    return "\n".join(lines)


def parse_md(content: bytes) -> str:
    import re as _re
    text = parse_txt(content)
    # Strip markdown syntax but keep readable text
    text = _re.sub(r'!\[([^\]]*)\]\([^)]+\)', r'\1', text)  # images
    text = _re.sub(r'\[([^\]]+)\]\([^)]+\)', r'\1', text)   # links
    text = _re.sub(r'#{1,6}\s+', '', text)                   # headings
    text = _re.sub(r'\*{1,3}([^*]+)\*{1,3}', r'\1', text)   # bold/italic
    text = _re.sub(r'`{1,3}[^`]*`{1,3}', '', text)          # code
    text = _re.sub(r'^[-*+]\s+', '', text, flags=_re.MULTILINE)  # list markers
    text = _re.sub(r'^\d+\.\s+', '', text, flags=_re.MULTILINE)  # numbered lists
    text = _re.sub(r'^>\s+', '', text, flags=_re.MULTILINE)      # blockquotes
    text = _re.sub(r'---+|===+', '', text)                   # horizontal rules
    return text


PARSERS = {
    ".txt": parse_txt,
    ".md": parse_md,
    ".markdown": parse_md,
    ".docx": parse_docx,
    ".pdf": parse_pdf,
    ".xlsx": parse_xlsx,
}


def parse_file(filename: str, content: bytes) -> str:
    ext = ""
    if "." in filename:
        ext = "." + filename.rsplit(".", 1)[1].lower()

    parser = PARSERS.get(ext)
    if not parser:
        raise ValueError(f"Nieobslugiwany format pliku: {ext}. Obslugiwane: {', '.join(PARSERS.keys())}")

    text = parser(content)
    return text[:15000]
